"""Tests del histórico propio de gestión (marcar / desmarcar / listar).

Usa la Postgres real del compose (fixtures `client`/`db_session` del conftest,
transacción revertida al final). No inventa datos de fuentes: reutiliza la
fuente ya sembrada (get-or-create por `codigo`) y crea convocatorias mínimas de
prueba con un `id_externo` propio del test.

Ciclo cubierto: marcar postulada -> desaparece del listado por defecto ->
aparece en GET /gestion -> `incluir_gestionadas=true` la devuelve con su
`estado_gestion` -> DELETE la reintegra al listado.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.convocatoria import Convocatoria
from app.models.fuente import Fuente


def _fuente(db: Session) -> Fuente:
    """Reutiliza la fuente `secop` ya sembrada; la crea si no existe."""
    fuente = db.execute(
        select(Fuente).where(Fuente.codigo == "secop")
    ).scalar_one_or_none()
    if fuente is None:
        fuente = Fuente(
            codigo="secop",
            nombre="SECOP II",
            url_base="https://example.org",
            tipo="api",
            activa=True,
            config={},
        )
        db.add(fuente)
        db.flush()
    return fuente


def _crear_convocatoria(db: Session, *, sufijo: str = "1", **campos) -> Convocatoria:
    fuente = _fuente(db)
    datos = dict(
        fuente_id=fuente.id,
        id_externo=f"TEST-GESTION-{sufijo}",
        hash_dedupe=f"test-gestion-dedupe-{sufijo}".ljust(64, "0"),
        hash_contenido=f"test-gestion-cont-{sufijo}".ljust(64, "0"),
        titulo=f"Convocatoria de prueba gestion {sufijo}",
        tipo="subvencion",
        estado="abierta",
        pais="Colombia",
        url_original=f"https://example.org/test-gestion-{sufijo}",
        keywords_match=[],
        raw={},
    )
    datos.update(campos)
    conv = Convocatoria(**datos)
    db.add(conv)
    db.flush()
    return conv


def _ids_del_listado(client, **params) -> list[int]:
    """Ids que devuelve `GET /convocatorias` con los filtros dados."""
    res = client.get("/api/v1/convocatorias", params={"page_size": 100, **params})
    assert res.status_code == 200, res.text
    return [item["id"] for item in res.json()["items"]]


def test_ciclo_completo_marcar_listar_desmarcar(client, db_session) -> None:
    conv = _crear_convocatoria(db_session, sufijo="ciclo")

    # 1. Antes de marcar: aparece en el listado y sin estado_gestion.
    assert conv.id in _ids_del_listado(client)

    # 2. Marcar como postulada.
    res = client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={"estado_gestion": "postulada", "responsable": "Ana", "notas": "enviada"},
    )
    assert res.status_code == 200, res.text
    marca = res.json()
    assert marca["convocatoria_id"] == conv.id
    assert marca["estado_gestion"] == "postulada"
    assert marca["responsable"] == "Ana"
    assert marca["notas"] == "enviada"
    # Sin fecha_postulacion en el request -> se usa el momento actual (no null).
    assert marca["fecha_postulacion"] is not None

    # 3. Desaparece del listado por defecto.
    assert conv.id not in _ids_del_listado(client)

    # 4. Aparece en el histórico con la convocatoria embebida.
    res = client.get("/api/v1/gestion", params={"page_size": 100})
    assert res.status_code == 200, res.text
    page = res.json()
    entrada = next(i for i in page["items"] if i["convocatoria_id"] == conv.id)
    assert entrada["estado_gestion"] == "postulada"
    assert entrada["convocatoria"]["id"] == conv.id
    assert entrada["convocatoria"]["estado_gestion"] == "postulada"
    assert entrada["convocatoria"]["url_original"] == conv.url_original

    # 5. incluir_gestionadas=true la devuelve, con su estado_gestion.
    res = client.get(
        "/api/v1/convocatorias",
        params={"page_size": 100, "incluir_gestionadas": "true"},
    )
    assert res.status_code == 200
    item = next(i for i in res.json()["items"] if i["id"] == conv.id)
    assert item["estado_gestion"] == "postulada"

    # 6. Desmarcar -> vuelve al listado y sale del histórico.
    res = client.delete(f"/api/v1/convocatorias/{conv.id}/gestion")
    assert res.status_code == 204
    assert conv.id in _ids_del_listado(client)
    res = client.get("/api/v1/gestion", params={"page_size": 100})
    assert all(i["convocatoria_id"] != conv.id for i in res.json()["items"])


def test_descartada_no_fija_fecha_postulacion(client, db_session) -> None:
    conv = _crear_convocatoria(db_session, sufijo="descartada")

    res = client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={"estado_gestion": "descartada", "fecha_postulacion": "2026-07-01T10:00:00Z"},
    )
    assert res.status_code == 200, res.text
    # Regla: `descartada` nunca lleva fecha de postulación (aunque venga en el body).
    assert res.json()["fecha_postulacion"] is None
    # También sale de la búsqueda.
    assert conv.id not in _ids_del_listado(client)


def test_fecha_postulacion_explicita_se_respeta(client, db_session) -> None:
    conv = _crear_convocatoria(db_session, sufijo="fecha")

    res = client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={
            "estado_gestion": "postulada",
            "fecha_postulacion": "2026-07-01T10:00:00Z",
        },
    )
    assert res.status_code == 200, res.text
    assert res.json()["fecha_postulacion"].startswith("2026-07-01T10:00:00")


def test_marcar_es_upsert(client, db_session) -> None:
    conv = _crear_convocatoria(db_session, sufijo="upsert")

    primera = client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={"estado_gestion": "postulada", "responsable": "Ana"},
    ).json()
    segunda = client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={"estado_gestion": "descartada", "responsable": "Beto"},
    ).json()

    # Mismo registro (UNIQUE por convocatoria), datos actualizados.
    assert segunda["id"] == primera["id"]
    assert segunda["estado_gestion"] == "descartada"
    assert segunda["responsable"] == "Beto"
    assert segunda["fecha_postulacion"] is None

    res = client.get("/api/v1/gestion", params={"page_size": 100})
    entradas = [i for i in res.json()["items"] if i["convocatoria_id"] == conv.id]
    assert len(entradas) == 1


def test_filtros_del_historico(client, db_session) -> None:
    postulada = _crear_convocatoria(db_session, sufijo="filtro-post")
    descartada = _crear_convocatoria(db_session, sufijo="filtro-desc")

    client.put(
        f"/api/v1/convocatorias/{postulada.id}/gestion",
        json={"estado_gestion": "postulada", "responsable": "Ana"},
    )
    client.put(
        f"/api/v1/convocatorias/{descartada.id}/gestion",
        json={"estado_gestion": "descartada", "responsable": "Beto"},
    )

    res = client.get(
        "/api/v1/gestion", params={"estado_gestion": "postulada", "page_size": 100}
    )
    ids = [i["convocatoria_id"] for i in res.json()["items"]]
    assert postulada.id in ids
    assert descartada.id not in ids

    res = client.get(
        "/api/v1/gestion", params={"responsable": "Beto", "page_size": 100}
    )
    ids = [i["convocatoria_id"] for i in res.json()["items"]]
    assert descartada.id in ids
    assert postulada.id not in ids

    # El filtro explícito en el listado implica incluir las gestionadas.
    ids = _ids_del_listado(client, estado_gestion="descartada")
    assert descartada.id in ids
    assert postulada.id not in ids


def test_filtros_ciudad_y_ambito(client, db_session) -> None:
    territorial = _crear_convocatoria(
        db_session,
        sufijo="ambito-terr",
        ciudad="Medellín",
        departamento="Antioquia",
        ambito="territorial",
    )
    nacional = _crear_convocatoria(
        db_session, sufijo="ambito-nac", ciudad="Bogotá", ambito="nacional"
    )

    ids = _ids_del_listado(client, ambito="territorial")
    assert territorial.id in ids
    assert nacional.id not in ids

    ids = _ids_del_listado(client, ciudad="Medellín")
    assert territorial.id in ids
    assert nacional.id not in ids

    # Ámbito inválido -> 422 (enum canónico).
    assert client.get("/api/v1/convocatorias", params={"ambito": "galactico"}).status_code == 422


def test_en_seguimiento_sigue_visible_en_busqueda(client, db_session) -> None:
    """A diferencia de postulada/descartada, `en_seguimiento` NO se oculta."""
    conv = _crear_convocatoria(db_session, sufijo="seguimiento-visible")

    res = client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={"estado_gestion": "en_seguimiento", "responsable": "Ana"},
    )
    assert res.status_code == 200, res.text
    marca = res.json()
    assert marca["estado_gestion"] == "en_seguimiento"
    # No se aplicó -> no lleva fecha de postulación.
    assert marca["fecha_postulacion"] is None

    # Sigue apareciendo en el listado por defecto, con su estado_gestion.
    listado = client.get("/api/v1/convocatorias", params={"page_size": 100}).json()
    item = next((i for i in listado["items"] if i["id"] == conv.id), None)
    assert item is not None, "en_seguimiento debe seguir en la búsqueda"
    assert item["estado_gestion"] == "en_seguimiento"

    # En cambio, al pasarla a postulada SÍ desaparece del listado por defecto.
    client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={"estado_gestion": "postulada", "responsable": "Ana"},
    )
    assert conv.id not in _ids_del_listado(client)


def test_stats_cuenta_aplicadas_y_en_seguimiento(client, db_session) -> None:
    base = client.get("/api/v1/stats").json()

    postulada = _crear_convocatoria(db_session, sufijo="stats-post")
    seguimiento = _crear_convocatoria(db_session, sufijo="stats-seg")
    client.put(
        f"/api/v1/convocatorias/{postulada.id}/gestion",
        json={"estado_gestion": "postulada", "responsable": "Ana"},
    )
    client.put(
        f"/api/v1/convocatorias/{seguimiento.id}/gestion",
        json={"estado_gestion": "en_seguimiento", "responsable": "Ana"},
    )

    stats = client.get("/api/v1/stats").json()
    assert stats["aplicadas"] == base["aplicadas"] + 1
    assert stats["en_seguimiento"] == base["en_seguimiento"] + 1


def test_marcar_convocatoria_inexistente_es_404(client) -> None:
    res = client.put(
        "/api/v1/convocatorias/999999/gestion", json={"estado_gestion": "postulada"}
    )
    assert res.status_code == 404


def test_desmarcar_sin_marca_es_404(client, db_session) -> None:
    conv = _crear_convocatoria(db_session, sufijo="sin-marca")
    assert client.delete(f"/api/v1/convocatorias/{conv.id}/gestion").status_code == 404
    assert client.delete("/api/v1/convocatorias/999999/gestion").status_code == 404


def test_estado_gestion_invalido_es_422(client, db_session) -> None:
    conv = _crear_convocatoria(db_session, sufijo="invalido")
    res = client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={"estado_gestion": "ganada"},
    )
    assert res.status_code == 422


def test_detalle_incluye_estado_gestion(client, db_session) -> None:
    conv = _crear_convocatoria(db_session, sufijo="detalle")
    assert client.get(f"/api/v1/convocatorias/{conv.id}").json()["estado_gestion"] is None

    client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={"estado_gestion": "postulada"},
    )
    detalle = client.get(f"/api/v1/convocatorias/{conv.id}").json()
    assert detalle["estado_gestion"] == "postulada"


def test_export_incluye_columnas_de_gestion(client, db_session) -> None:
    conv = _crear_convocatoria(db_session, sufijo="export")
    client.put(
        f"/api/v1/convocatorias/{conv.id}/gestion",
        json={
            "estado_gestion": "postulada",
            "responsable": "Ana",
            "fecha_postulacion": "2026-07-01T10:00:00Z",
        },
    )

    res = client.post("/api/v1/convocatorias/export", json={"ids": [conv.id]})
    assert res.status_code == 200
    ws = load_workbook(BytesIO(res.content)).active
    encabezados = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    assert "Estado de gestión" in encabezados
    assert "Responsable" in encabezados
    assert "Fecha de postulación" in encabezados

    fila = {
        encabezados[i - 1]: ws.cell(row=2, column=i).value
        for i in range(1, ws.max_column + 1)
    }
    assert fila["Estado de gestión"] == "postulada"
    assert fila["Responsable"] == "Ana"
    assert fila["Fecha de postulación"] == "2026-07-01"
