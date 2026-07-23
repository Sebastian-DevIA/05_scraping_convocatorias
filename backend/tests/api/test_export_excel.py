"""Tests de `POST /convocatorias/export` (descarga a Excel de seleccionadas).

Usa la Postgres real del compose (fixtures `client`/`db_session` del conftest,
transacción revertida al final). No inventa datos de fuentes: reutiliza una
fuente ya sembrada (o la crea si falta) y añade una convocatoria mínima de
prueba, luego verifica que el `.xlsx` la contiene.

Nota de aislamiento: los tests corren contra la MISMA Postgres del compose, cuyo
`entrypoint.sh` ya sembró las fuentes reales (commiteadas fuera de la transacción
del test). Por eso NO se crea una fuente con un `codigo` que ya existe: se
reutiliza la sembrada por `codigo` (get-or-create).
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.convocatoria import Convocatoria
from app.models.fuente import Fuente


def _fuente(db: Session) -> Fuente:
    """Reutiliza la fuente `secop` ya sembrada; la crea si no existe."""
    fuente = db.execute(
        select(Fuente).where(Fuente.codigo == "secop")
    ).scalar_one_or_none()
    if fuente is None:
        fuente = Fuente(
            codigo="secop",
            nombre="SECOP II",
            url_base="https://example.org",
            tipo="api",
            activa=True,
            config={},
        )
        db.add(fuente)
        db.flush()
    return fuente


def _crear_convocatoria(db: Session, *, sufijo: str = "1") -> Convocatoria:
    fuente = _fuente(db)
    conv = Convocatoria(
        fuente_id=fuente.id,
        id_externo=f"TEST-EXPORT-{sufijo}",
        hash_dedupe=f"test-export-dedupe-{sufijo}".ljust(64, "0"),
        hash_contenido=f"test-export-cont-{sufijo}".ljust(64, "0"),
        titulo="Fondo semilla para nuevas fundaciones",
        descripcion="Linea 1\nLinea 2",
        entidad="Fundación X",
        tipo="subvencion",
        estado="abierta",
        pais="Colombia",
        url_original=f"https://example.org/test-export-{sufijo}",
        keywords_match=[],
        apto_fundaciones_nuevas=True,
        raw={},
    )
    db.add(conv)
    db.flush()
    return conv


def test_export_xlsx_ok(client, db_session) -> None:
    conv = _crear_convocatoria(db_session)

    res = client.post("/api/v1/convocatorias/export", json={"ids": [conv.id]})

    assert res.status_code == 200
    assert res.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disp = res.headers.get("content-disposition", "")
    assert "attachment" in disp
    assert "convocatorias_seleccionadas.xlsx" in disp

    wb = load_workbook(BytesIO(res.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Título"  # encabezado
    valores_fila = [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)]
    assert "Fondo semilla para nuevas fundaciones" in valores_fila
    assert "https://example.org/test-export-1" in valores_fila
    assert "Sí" in valores_fila  # apto_fundaciones_nuevas -> "Sí"
    # Los saltos de línea reales se preservan dentro de la celda (Excel multilínea).
    assert "Linea 1\nLinea 2" in valores_fila


def test_export_mezcla_existentes_e_inexistentes(client, db_session) -> None:
    conv = _crear_convocatoria(db_session, sufijo="mix")
    res = client.post(
        "/api/v1/convocatorias/export", json={"ids": [conv.id, 999999]}
    )
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb.active
    # Solo la existente entra; la inexistente se ignora -> 1 encabezado + 1 fila.
    assert ws.max_row == 2


def test_export_ids_vacios_es_422(client) -> None:
    res = client.post("/api/v1/convocatorias/export", json={"ids": []})
    assert res.status_code == 422


def test_export_ids_inexistentes_devuelve_xlsx_vacio(client) -> None:
    # Ids que no existen: se ignoran; el archivo trae solo encabezados.
    res = client.post("/api/v1/convocatorias/export", json={"ids": [999999]})
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Título"
    assert ws.max_row == 1  # solo la fila de encabezados
