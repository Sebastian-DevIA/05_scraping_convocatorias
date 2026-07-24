"""Lógica de consulta de convocatorias: listado con filtros y detalle.

La forma exacta de la respuesta está congelada en `docs/api-contract.md`.
`fuente_codigo`/`fuente_nombre` se componen aquí a partir del join con `fuentes`
(joinedload, sin N+1). El full-text usa la columna generada `busqueda`
(tsvector español) con `websearch_to_tsquery('spanish', q)`.
"""

from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO

from sqlalchemy import ColumnElement, func, select
from sqlalchemy.orm import Session, joinedload

from app.api.deps import ConvocatoriaFiltros, PaginacionParams
from app.constants import ESTADOS_GESTION_OCULTAN
from app.models.convocatoria import Convocatoria
from app.models.fuente import Fuente
from app.models.gestion import Gestion
from app.schemas.convocatoria import (
    ConvocatoriaDetailResponse,
    ConvocatoriaPageResponse,
    ConvocatoriaResponse,
)

# Campo de orden admitido -> columna real de la tabla.
_ORDEN_COLUMNAS = {
    "fecha_publicacion": Convocatoria.fecha_publicacion,
    "fecha_cierre": Convocatoria.fecha_cierre,
    "monto": Convocatoria.monto,
    "ultima_vez_visto": Convocatoria.ultima_vez_visto,
}


def _inicio_dia(dia: date) -> datetime:
    """Medianoche UTC del día dado (los timestamps de dominio son TIMESTAMPTZ)."""
    return datetime.combine(dia, time.min, tzinfo=timezone.utc)


def _existe_gestion(*extra: ColumnElement[bool]) -> ColumnElement[bool]:
    """EXISTS correlacionado contra `gestiones` (0 o 1 registro por convocatoria).

    Se usa EXISTS / NOT EXISTS (no un LEFT JOIN con filtro) para que el conteo
    de la paginación siga contando filas de `convocatorias`.
    """
    return (
        select(Gestion.id)
        .where(Gestion.convocatoria_id == Convocatoria.id, *extra)
        .exists()
    )


def _construir_condiciones(f: ConvocatoriaFiltros) -> list[ColumnElement[bool]]:
    """Traduce los filtros del contrato a condiciones WHERE de SQLAlchemy."""
    condiciones: list[ColumnElement[bool]] = []

    if f.q:
        # Full-text en español contra la columna generada `busqueda` (tsvector).
        condiciones.append(
            Convocatoria.busqueda.bool_op("@@")(func.websearch_to_tsquery("spanish", f.q))
        )
    if f.fuente:
        # Código inexistente -> subquery vacía -> 0 resultados (sin 404 en el listado).
        condiciones.append(
            Convocatoria.fuente_id.in_(select(Fuente.id).where(Fuente.codigo == f.fuente))
        )
    if f.estado:
        condiciones.append(Convocatoria.estado == f.estado)
    if f.tipo:
        condiciones.append(Convocatoria.tipo == f.tipo)
    if f.departamento:
        condiciones.append(Convocatoria.departamento == f.departamento)
    if f.ciudad:
        condiciones.append(Convocatoria.ciudad == f.ciudad)
    if f.ambito:
        condiciones.append(Convocatoria.ambito == f.ambito)
    if f.apto_fundaciones_nuevas is not None:
        condiciones.append(
            Convocatoria.apto_fundaciones_nuevas.is_(f.apto_fundaciones_nuevas)
        )

    if f.estado_gestion:
        # Filtro explícito por el histórico: implica incluir las gestionadas.
        condiciones.append(_existe_gestion(Gestion.estado_gestion == f.estado_gestion))
    elif not f.incluir_gestionadas:
        # Default: se ocultan solo las `postulada`/`descartada` (para no repetir
        # una postulación). Las `en_seguimiento` siguen visibles: aún no se aplicó.
        condiciones.append(
            ~_existe_gestion(Gestion.estado_gestion.in_(ESTADOS_GESTION_OCULTAN))
        )

    if f.fecha_publicacion_desde:
        condiciones.append(Convocatoria.fecha_publicacion >= _inicio_dia(f.fecha_publicacion_desde))
    if f.fecha_publicacion_hasta:
        # `hasta` inclusivo del día completo: < medianoche del día siguiente.
        condiciones.append(
            Convocatoria.fecha_publicacion < _inicio_dia(f.fecha_publicacion_hasta + timedelta(days=1))
        )
    if f.fecha_cierre_desde:
        condiciones.append(Convocatoria.fecha_cierre >= _inicio_dia(f.fecha_cierre_desde))
    if f.fecha_cierre_hasta:
        condiciones.append(
            Convocatoria.fecha_cierre < _inicio_dia(f.fecha_cierre_hasta + timedelta(days=1))
        )

    if f.monto_min is not None:
        condiciones.append(Convocatoria.monto >= f.monto_min)
    if f.monto_max is not None:
        condiciones.append(Convocatoria.monto <= f.monto_max)

    return condiciones


def _order_by(orden: str) -> list:
    """`orden` validado por el patrón del contrato. `-campo` = descendente.

    Se añade `id DESC` como desempate para una paginación estable.
    """
    descendente = orden.startswith("-")
    campo = orden[1:] if descendente else orden
    columna = _ORDEN_COLUMNAS[campo]
    principal = columna.desc() if descendente else columna.asc()
    return [principal, Convocatoria.id.desc()]


def a_convocatoria_response(c: Convocatoria) -> ConvocatoriaResponse:
    """Modelo ORM (con `fuente` y `gestion` cargadas) -> schema de listado.

    Público porque el servicio de gestión embebe la convocatoria en su
    histórico y debe producir exactamente la misma forma.
    """
    return ConvocatoriaResponse(
        id=c.id,
        id_externo=c.id_externo,
        fuente_id=c.fuente_id,
        fuente_codigo=c.fuente.codigo,
        fuente_nombre=c.fuente.nombre,
        titulo=c.titulo,
        descripcion=c.descripcion,
        entidad=c.entidad,
        tipo=c.tipo,
        modalidad=c.modalidad,
        estado=c.estado,
        monto=c.monto,
        moneda=c.moneda,
        departamento=c.departamento,
        ciudad=c.ciudad,
        pais=c.pais,
        ambito=c.ambito,
        fecha_publicacion=c.fecha_publicacion,
        fecha_apertura=c.fecha_apertura,
        fecha_cierre=c.fecha_cierre,
        url_original=c.url_original,
        keywords_match=list(c.keywords_match or []),
        apto_fundaciones_nuevas=c.apto_fundaciones_nuevas,
        # None si nunca se marcó (dato NUESTRO, no de la fuente).
        estado_gestion=c.gestion.estado_gestion if c.gestion else None,
        primera_vez_visto=c.primera_vez_visto,
        ultima_vez_visto=c.ultima_vez_visto,
        creado_en=c.creado_en,
        actualizado_en=c.actualizado_en,
    )


def _a_detail_response(c: Convocatoria, include_raw: bool) -> ConvocatoriaDetailResponse:
    """Añade `requisitos` y `raw` (solo si `include_raw`) al schema de detalle."""
    base = a_convocatoria_response(c)
    return ConvocatoriaDetailResponse(
        **base.model_dump(),
        requisitos=c.requisitos,
        raw=c.raw if include_raw else None,
    )


def listar_convocatorias(
    db: Session, filtros: ConvocatoriaFiltros, paginacion: PaginacionParams
) -> ConvocatoriaPageResponse:
    """Listado paginado con filtros -> `{items, total, page, page_size}`."""
    condiciones = _construir_condiciones(filtros)

    total = db.execute(
        select(func.count()).select_from(Convocatoria).where(*condiciones)
    ).scalar_one()

    filas = (
        db.execute(
            select(Convocatoria)
            # `gestion` es 0..1 (UNIQUE) -> el joinedload no multiplica filas y
            # el OFFSET/LIMIT sigue siendo por convocatoria.
            .options(joinedload(Convocatoria.fuente), joinedload(Convocatoria.gestion))
            .where(*condiciones)
            .order_by(*_order_by(filtros.orden))
            .offset(paginacion.offset)
            .limit(paginacion.limit)
        )
        .scalars()
        .all()
    )

    return ConvocatoriaPageResponse(
        items=[a_convocatoria_response(c) for c in filas],
        total=total,
        page=paginacion.page,
        page_size=paginacion.page_size,
    )


def obtener_convocatoria(
    db: Session, convocatoria_id: int, include_raw: bool
) -> ConvocatoriaDetailResponse | None:
    """Detalle por id (con `requisitos`, y `raw` si `include_raw`). None si no existe."""
    convocatoria = db.execute(
        select(Convocatoria)
        .options(joinedload(Convocatoria.fuente), joinedload(Convocatoria.gestion))
        .where(Convocatoria.id == convocatoria_id)
    ).scalar_one_or_none()

    if convocatoria is None:
        return None
    return _a_detail_response(convocatoria, include_raw)


def _fecha_excel(dt: datetime | None) -> str:
    """Fecha en formato YYYY-MM-DD para el Excel. None -> cadena vacía."""
    return dt.strftime("%Y-%m-%d") if dt else ""


# Columnas del Excel: (encabezado, ancho, extractor). El orden prioriza lo que
# se necesita para DECIDIR y PARTICIPAR; `url_original` permite verificar que la
# convocatoria existe en la fuente oficial.
_COLUMNAS_EXPORT = [
    ("Título", 42, lambda c: c.titulo),
    ("Entidad emisora", 32, lambda c: c.entidad or ""),
    ("Fuente", 24, lambda c: c.fuente.nombre),
    ("Estado", 12, lambda c: c.estado),
    ("Tipo", 12, lambda c: c.tipo),
    ("Modalidad", 20, lambda c: c.modalidad or ""),
    ("País", 16, lambda c: c.pais or ""),
    ("Departamento", 16, lambda c: c.departamento or ""),
    ("Ciudad", 16, lambda c: c.ciudad or ""),
    ("Monto", 16, lambda c: float(c.monto) if c.monto is not None else ""),
    ("Moneda", 8, lambda c: c.moneda or ""),
    ("Publicación", 13, lambda c: _fecha_excel(c.fecha_publicacion)),
    ("Apertura", 13, lambda c: _fecha_excel(c.fecha_apertura)),
    ("Cierre", 13, lambda c: _fecha_excel(c.fecha_cierre)),
    ("Apta fundaciones nuevas", 14, lambda c: "Sí" if c.apto_fundaciones_nuevas else "No"),
    # Histórico propio (dato NUESTRO, vacío si la convocatoria no se ha gestionado).
    ("Estado de gestión", 16, lambda c: c.gestion.estado_gestion if c.gestion else ""),
    ("Responsable", 20, lambda c: (c.gestion.responsable or "") if c.gestion else ""),
    (
        "Fecha de postulación",
        15,
        lambda c: _fecha_excel(c.gestion.fecha_postulacion) if c.gestion else "",
    ),
    ("Requisitos", 45, lambda c: c.requisitos or ""),
    ("Descripción", 60, lambda c: c.descripcion or ""),
    ("Palabras clave", 24, lambda c: ", ".join(c.keywords_match or [])),
    ("URL original (verificar)", 55, lambda c: c.url_original),
    ("ID externo", 22, lambda c: c.id_externo),
]


def exportar_convocatorias_excel(db: Session, ids: list[int]) -> bytes:
    """Genera un `.xlsx` (bytes) con las convocatorias seleccionadas.

    Incluye los datos necesarios para participar y `url_original` para verificar
    que la convocatoria existe en la fuente oficial. Los ids inexistentes se
    ignoran. Se ordena por fecha de cierre (próximas primero; sin fecha al final).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    filas = (
        db.execute(
            select(Convocatoria)
            .options(joinedload(Convocatoria.fuente), joinedload(Convocatoria.gestion))
            .where(Convocatoria.id.in_(ids))
            .order_by(
                Convocatoria.fecha_cierre.asc().nulls_last(),
                Convocatoria.id.asc(),
            )
        )
        .scalars()
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Convocatorias"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_align = Alignment(vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)

    for col_idx, (header, width, _fn) in enumerate(_COLUMNAS_EXPORT, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx, convocatoria in enumerate(filas, start=2):
        for col_idx, (_header, _width, extractor) in enumerate(_COLUMNAS_EXPORT, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=extractor(convocatoria))
            cell.alignment = cell_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNAS_EXPORT))}{len(filas) + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
